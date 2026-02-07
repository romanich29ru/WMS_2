from openpyxl import load_workbook
from collections import Counter, defaultdict
from pathlib import Path

file_path = Path('Занято-Свободно2.xlsx')
if not file_path.exists():
    raise FileNotFoundError(file_path)

wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
print('sheets =', wb.sheetnames)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print('empty sheet')
    wb.close()
    raise SystemExit

header = rows[0]
print('\nHeader with indices:')
for i, h in enumerate(header, start=1):
    print(i, repr(h))

# Show counts of non-empty values per column in first 500 rows
col_nonempty = defaultdict(int)
col_samples = defaultdict(list)
for rnum, row in enumerate(rows[1:501], start=2):
    for i, val in enumerate(row, start=1):
        if val is not None and str(val).strip() != '':
            col_nonempty[i] += 1
            if len(col_samples[i]) < 5:
                col_samples[i].append(val)

print('\nNon-empty counts per column (first 500 data rows):')
for i in sorted(col_nonempty.keys()):
    print(i, col_nonempty[i], 'samples=', col_samples[i])

print('\nFirst 50 full rows:')
for rnum, row in enumerate(rows[1:51], start=2):
    print(rnum, row)

# For header names we expect: Название Ячейки, Статус, and possibly Артикулы
# Find column index by header name search
name_candidates = [i+1 for i,h in enumerate(header) if h and 'Название' in str(h)]
status_candidates = [i+1 for i,h in enumerate(header) if h and 'Статус' in str(h)]
print('\nDetected header columns:')
print('name_candidates=', name_candidates)
print('status_candidates=', status_candidates)

# Heuristic: find columns where values look like article lists (contain commas or spaces with digits/letters)
article_like = []
for i in range(1, len(header)+1):
    samples = col_samples.get(i, [])
    score = 0
    for s in samples:
        sstr = str(s)
        if ',' in sstr or ';' in sstr or (' ' in sstr and any(ch.isdigit() for ch in sstr)) or (len(sstr.split()) > 1 and any(len(tok) >= 3 for tok in sstr.split())):
            score += 1
    if score > 0:
        article_like.append((i, score, samples))

print('\nLikely article columns (index,score,samples):')
for item in article_like:
    print(item)

wb.close()
