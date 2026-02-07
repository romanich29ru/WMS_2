from openpyxl import load_workbook
from collections import Counter
from pathlib import Path

file_path = Path('Занято-Свободно2.xlsx')
if not file_path.exists():
    raise FileNotFoundError(file_path)

wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
print('sheets =', wb.sheetnames)
ws = wb.active

iter_rows = ws.iter_rows(values_only=True)

# Read header
try:
    header = next(iter_rows)
except StopIteration:
    header = None

print('header_row (len={}):'.format(len(header) if header else 0))
print(header)

# We'll inspect columns C(3), D(4), I(9) as 1-based indexes -> 2,3,8 zero-based
c_idx = 2
d_idx = 3
i_idx = 8

first_rows = []
status_counter = Counter()
articles_counter = Counter()
cellname_counter = Counter()

for rnum, row in enumerate(iter_rows, start=2):
    # row can be shorter than expected
    val_c = row[c_idx] if len(row) > c_idx else None
    val_d = row[d_idx] if len(row) > d_idx else None
    val_i = row[i_idx] if len(row) > i_idx else None

    first_rows.append((rnum, val_c, val_d, val_i, row))
    status_counter.update([str(val_d).strip() if val_d is not None else ''])
    articles_counter.update([str(val_i).strip() if val_i is not None else ''])
    cellname_counter.update([str(val_c).strip() if val_c is not None else ''])

    if rnum >= 21:
        break

print('\nFirst 20 data rows (Row#, C, D, I):')
for rnum, val_c, val_d, val_i, row in first_rows:
    print(rnum, '|', repr(val_c), '|', repr(val_d), '|', repr(val_i))

# Print distinct non-empty statuses and counts
print('\nDistinct status values (column D) with counts:')
for status, cnt in status_counter.most_common():
    if status and status != 'None' and status.lower()!='none':
        print(f'{repr(status)}: {cnt}')

# Show sample articles values (column I) with counts (top 20):
print('\nSample articles values (column I) with counts (top 20):')
for art, cnt in articles_counter.most_common(20):
    if art and art.strip():
        print(f'{repr(art)}: {cnt}')

# Detect anomalies in cell names
print("\nSample cell name anomalies (presence of /, \\\\, multiple spaces, non-ascii):")
for name, cnt in cellname_counter.most_common(30):
    if not name:
        continue
    issues = []
    if '/' in name:
        issues.append('/')
    if '\\' in name:
        issues.append('\\')
    if '  ' in name:
        issues.append('double-space')
    if any(ord(ch) > 127 for ch in name):
        issues.append('non-ascii')
    if issues:
        print(f'{repr(name)} ({cnt}) ->', ','.join(issues))

wb.close()
