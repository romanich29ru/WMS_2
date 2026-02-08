import json
import re
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

file_path = Path('Занято-Свободно2.xlsx')
if not file_path.exists():
    raise FileNotFoundError(file_path)

wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
if not rows:
    print('empty sheet')
    wb.close()
    raise SystemExit(1)

header = [h if h is not None else '' for h in rows[0]]
# Normalize headers for matching
hdr_norm = [str(h).strip().lower() for h in header]

# Heuristics for columns
def find_header_candidates(keywords):
    keys = []
    for i, h in enumerate(hdr_norm):
        for kw in keywords:
            if kw in h:
                keys.append(i)
                break
    return keys

name_candidates = find_header_candidates(['назван', 'ячейк', 'ячейка', 'cell', 'address'])
status_candidates = find_header_candidates(['статус', 'status'])

# scan first N rows to detect which columns contain known status words
status_words = set(['свободна', 'занято', 'занята', 'occupied', 'empty', 'empty ', 'vacant'])
col_status_scores = Counter()
col_articles_scores = Counter()

N = min(500, max(50, len(rows)-1))
for row in rows[1:1+N]:
    for i, val in enumerate(row):
        if val is None:
            continue
        s = str(val).strip()
        sl = s.lower()
        # status detection
        for w in status_words:
            if w in sl:
                col_status_scores[i] += 1
        # article-like detection
        # stronger signal when values contain separators (comma/semicolon) or tokens that look like SKUs (digits, underscores)
        if ',' in s or ';' in s:
            col_articles_scores[i] += 3
        else:
            toks = re.split(r"[\s,;]+", s)
            sku_score = 0
            addr_score = 0
            for t in toks:
                t = t.strip()
                if not t:
                    continue
                # token looks like SKU: mostly digits, or digits+underscore, or long alnum
                if re.fullmatch(r"[0-9_\-]{3,}", t) or re.fullmatch(r"[A-Z0-9_]{5,}", t, flags=re.IGNORECASE):
                    sku_score += 1
                # token looks like address fragment (short alpha parts like 'OS' or 'NA')
                if re.fullmatch(r"[A-Za-z]{1,3}", t):
                    addr_score += 1
            if sku_score > 0 and sku_score >= addr_score:
                col_articles_scores[i] += sku_score

# pick best columns
status_col = None
if status_candidates:
    status_col = status_candidates[0]
elif col_status_scores:
    status_col = col_status_scores.most_common(1)[0][0]

articles_col = None
if col_articles_scores:
    # prefer article columns that are not detected as name/status columns
    for idx, score in col_articles_scores.most_common():
        if idx in name_candidates:
            continue
        if status_candidates and idx in status_candidates:
            continue
        if status_col is not None and idx == status_col:
            continue
        articles_col = idx
        break

# fallback: if header length >= 9, use 3(C),4(D),9(I) as user expected indexes
if status_col is None and len(header) >= 4:
    status_col = 3  # 0-based index for column 4
if articles_col is None and len(header) >= 9:
    articles_col = 8  # 0-based index for column 9

print('Detected columns -> name_candidates=', name_candidates, 'status_col=', status_col, 'articles_col=', articles_col)

# function to split articles into tokens
def split_articles_field(val):
    if val is None:
        return []
    s = str(val).strip()
    if s == '':
        return []
    # prefer splitting by comma/semicolon
    if ',' in s or ';' in s:
        parts = re.split(r'[;,]', s)
    else:
        parts = re.split(r'\s+', s)
    # cleanup
    parts = [p.strip() for p in parts if p and len(p.strip())>=1]
    # filter out very short tokens (1 char) unless they include digits
    parts = [p for p in parts if len(p) > 1 or re.search(r"\d", p)]
    return parts

# parse rows and build preview
preview = []
counts = Counter()
for rnum, row in enumerate(rows[1:], start=2):
    # get name
    name = None
    if name_candidates:
        name = row[name_candidates[0]] if len(row) > name_candidates[0] else None
    else:
        name = row[2] if len(row) > 2 else None
    status = row[status_col] if status_col is not None and len(row) > status_col else None
    articles_raw = row[articles_col] if articles_col is not None and len(row) > articles_col else None
    articles = split_articles_field(articles_raw)
    articles_count = len(articles)
    needs_attention = articles_count >= 2
    parsed_tokens = []
    if name is not None:
        parsed_tokens = re.split(r'[\s\-_/]+', str(name).strip())
    preview.append({
        'row': rnum,
        'cell_name_raw': name,
        'cell_name_tokens': parsed_tokens,
        'status_raw': status,
        'articles_raw': articles_raw,
        'articles_count': articles_count,
        'needsAttention': needs_attention
    })
    counts['total'] += 1
    if needs_attention:
        counts['needsAttention'] += 1
    if articles_count == 0:
        counts['noArticles'] += 1

# Save preview JSON
out = {
    'file': str(file_path),
    'detected': {
        'name_candidates': name_candidates,
        'status_col': status_col,
        'articles_col': articles_col
    },
    'summary': counts,
    'preview_rows': preview[:200]
}

out_path = Path('scripts') / 'wms_parsed_preview.json'
with out_path.open('w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print('Wrote preview to', out_path)
print('Summary:', counts)
wb.close()
